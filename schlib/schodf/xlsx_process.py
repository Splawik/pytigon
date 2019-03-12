from openpyxl import load_workbook, Workbook
import re


def can_append_to_data_table(wb):
    for sheet in wb._sheets:
        for table in sheet._tables:
            if table.name == 'data_table':
                return True
    return False


def append_to_data_table(wb, data_table):
    ws = None
    table = None
    for sheet in wb._sheets:
        for tbl in sheet._tables:
            if tbl.name == 'data_table':
                table = tbl
                ws = sheet
                break
        if ws:
            break

    if table and ws:
        x = table.ref.split(':')[-1]

        idx = re.sub("[^0-9]", "", x)
        for row in data_table:
            ws.append(row)

        ws.delete_rows(int(idx))

        end_idx = int(idx) + len(data_table) - 1
        table.ref = table.ref[:-1 * len(idx)] + str(end_idx)
        return True
    else:
        return False


# rec:
# [ "ws_name=>", rec]
# [ "=>", rec] (ws = data)
# [ rec ]

# ["ws_name//A1", "value"]
# ["ws_name//A1:value"]



def transform_xlsx(stream_in, stream_out, transform_list):

    first_ws = False
    if stream_in:
        wb = load_workbook(stream_in)
    else:
        wb = Workbook()
        first_ws = True

    if can_append_to_data_table(wb):
        dt = True
    else:
        dt = False

    data_table = []

    ws_name_prev = None
    ws = None

    def set_sheet(name):
        nonlocal ws_name_prev, wb, ws
        if name != ws_name_prev:
            if not ws_name_prev and first_ws:
                ws = wb.active
                ws.title = name
            try:
                ws = wb[name]
            except:
                ws = wb.create_sheet(name)
            ws_name_prev = name

    for _pos in transform_list:
        pos = list(_pos)
        key = pos[0]

        if type(key) == str and '=>' in key:
            x = key.split('=>')
            ws_name = x[0] if x[0] else 'data'
            set_sheet(ws_name)
            ws.append(pos[1:])

        elif type(key) == str and '//' in key:
            x = key.split('//')
            ws_name = x[0] if x[0] else 'data'
            set_sheet(ws_name)
            if ':' in x[1]:
                xx = x[1].split(':', 1)
            else:
                xx = (x[1], pos[1],)

            ws[xx[0]] = xx[1]

        else:
            if dt:
                data_table.append(pos)
            else:
                ws_name = 'data'
                set_sheet(ws_name)
                ws.append(pos)

        if data_table:
            append_to_data_table(wb, data_table)

    wb.save(stream_out)
