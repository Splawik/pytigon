#!/usr/bin/python

# -*- coding: utf-8 -*-
from django.http import HttpResponseRedirect, HttpResponse
from django.shortcuts import render, redirect
from django import forms
from django.template.loader import render_to_string
from django.template import Context, Template
from django.template import RequestContext
from django.conf import settings
from django.views.generic import TemplateView

from pytigon_lib.schviews.form_fun import form_with_perms
from pytigon_lib.schviews.viewtools import dict_to_template, dict_to_odf, dict_to_pdf, dict_to_json, dict_to_xml
from pytigon_lib.schviews.viewtools import render_to_response
from pytigon_lib.schdjangoext.tools import make_href

from django.utils.translation import ugettext_lazy as _

from . import models
import os
import sys
import datetime

from pytigon_lib.schfs.vfstools import get_temp_filename
from pytigon_lib.schdjangoext.tools import gettempdir

from pytigon_lib.schdjangoext.spreadsheet_render import render_to_response_odf
from pytigon_lib.schviews.viewtools import render_to_response_ext
import  openpyxl
from pytigon_lib.schdjangoext.spreadsheet_render import render_odf
import tempfile
from django.core.mail import send_mail, EmailMessage, EmailMultiAlternatives

import os, zipfile, io

select1 = """
select gniazdo, 
sum(H1)*60+sum(M1) as MIN_godziny_pracy, sum(H2)*60+sum(M2) as MIN_praca_netto, sum(H3)*60+sum(M3) as MIN_naprawa, sum(H4)*60+sum(M4) as MIN_konswerwacja, sum(H5)*60+sum(M5) as MIN_przestoj, 
sum(H6)*60+sum(M6) as MIN_zmiana_formy, sum(H7)*60+sum(M7) as MIN_brak_betonu, sum(H8)*60+sum(M8) as MIN_brak_palet, sum(H9)*60+sum(M9) as MIN_pelna_wieza,
sum(IloscWymianFormy) as IloscWymianFormy,
sum(CzasNormatywny_HH)*60+sum(CzasNormatywny_MM) as CzasNormatywny_MM,
sum(cast(IloscZmian as float)) as IloscZmian, 
sum(RP_01) as RP_01,  sum(RP_02) as RP_02, sum(RP_03) as RP_03,  sum(RP_04) as RP_04, sum(RP_05) as RP_05, sum(RP_06) as RP_06, sum(RP_07) as RP_07,
sum(CzasKalkulacyjny_HH)*60+sum(CzasKalkulacyjny_MM) as CzasKalkulacyjny_MM,
sum(CzasNormatywny) as CzasNormatywny
from
(
    select Nag.NagId, nag.gniazdo, 
    max(nag.PD_RP_H1) as H1, max(nag.PD_RP_M1) as M1, max(nag.PD_RP_H2) as H2, max(nag.PD_RP_M2) as M2, max(nag.PD_RP_H3) as H3, max(nag.PD_RP_M3) as M3, max(nag.PD_RP_H4) as H4, max(nag.PD_RP_M4) as M4, max(nag.PD_RP_H5) as H5, max(nag.PD_RP_M5) as M5, 
    max(nag.PD_RP_KomentarzDoPrzestoju) as KomentarzDoPrzestoju,
    max(nag.PD_RP_H6) as H6, max(nag.PD_RP_M6) as M6, max(nag.PD_RP_H7) as H7, max(nag.PD_RP_M7) as M7, max(nag.PD_RP_H8) as H8, max(nag.PD_RP_M8) as M8, max(nag.PD_RP_H9) as H9, max(nag.PD_RP_M9) as M9,
    max(nag.PD_IloscWymianFormy) as IloscWymianFormy,
    max(nag.CzasNormatywny_HH) as CzasNormatywny_HH, max(nag.CzasNormatywny_MM) as CzasNormatywny_MM,
    max(nag.pd_IloscZmian) as IloscZmian, max(nag.PD_RP_O1) as RP_01,  max(nag.PD_RP_O2) as RP_02,  max(nag.PD_RP_O3) as RP_03,  max(nag.PD_RP_O4) as RP_04,  max(nag.PD_RP_O5) as RP_05,  max(nag.PD_RP_O6) as RP_06,  max(nag.PD_RP_O7) as RP_07,
    max(nag.PD_CzasKalkulacyjny_HH) as CzasKalkulacyjny_HH, max(nag.PD_CzasKalkulacyjny_MM) as CzasKalkulacyjny_MM,
    max(nag.Opis) as Opis, max(nag.CzasNormatywny) as CzasNormatywny

    from dbo.w_mg_vv_Nag_RP_E nag(NOLOCK)
    join dbo.mg_vv_Lin_RejAll_E lin WITH (NOLOCK) on lin.NagId = nag.NagId

    where nag.Data >='%s' and nag.Data< '%s' and nag.rd = 'PD'
    group by nag.NagId, nag.gniazdo
) tab
group by gniazdo
"""

select2 = """
select nag.gniazdo, k.SymKar as Kar,
sum(lin.PD_IloscMGPrzel) as IloscMGPrzel, sum(lin.PD_IloscCykli) as IloscCykli, sum(lin.PD_IloscCykli*ccp.CzasCyklu)/sum(lin.PD_IloscCykli) as CzasCyklu, 
sum(lin.PD_IloscCykli*ccp.CzasCyklu) as CzasTeoretyczny,
sum( 
case when nag.CzasNormatywny_HH <> 0 or nag.CzasNormatywny_MM <> 0 then
(lin.PD_IloscCykli*ccp.CzasCyklu)*(nag.PD_CzasKalkulacyjny_HH*3600+nag.PD_CzasKalkulacyjny_MM*60)/(nag.CzasNormatywny_HH*3600+nag.CzasNormatywny_MM*60) 
else 0 end 
) as CzasKalkulacyjny

from dbo.w_mg_vv_Nag_RP_E nag(NOLOCK)
join dbo.mg_vv_Lin_RejAll_E lin WITH (NOLOCK) on lin.NagId = nag.NagId
INNER JOIN dbo.mg_Kar k WITH (NOLOCK) ON k.SymKar = lin.SymKar
LEFT JOIN dbo.w_mg_CzasyCykliProdKartotek ccp WITH (NOLOCK) ON ccp.SymKar = lin.SymKar AND ccp.Gniazdo = lin.Gniazdo AND ccp.DataOd <= nag.Data AND ccp.DataDo >= nag.Data

where nag.Data >='%s' and nag.Data< '%s' and nag.rd = 'PD' and nag.status=1
group by nag.gniazdo, k.SymKar
having sum(lin.PD_IloscCykli)>0
order by nag.gniazdo, CzasKalkulacyjny DESC
"""

select3 = """
select nag.Gniazdo, REPLACE(left(OpiKar4,1), 'P', '1') Kategoria, sum(PD_IloscMGPrzel) as IloscMGPrzel from 
mg_Nag nag(NOLOCK)
left join mg_Lin lin(NOLOCK) on lin.NagId = nag.NagId
left join mg_kar kar(NOLOCK) on kar.symkar = lin.symkar 
where nag.Data >='%s' and nag.Data< '%s' and nag.RD = 'PW' and OpiKar4 like '[1,2,3,P]%%' and nag.status > 0
group by lin.mag, nag.gniazdo, REPLACE(left(OpiKar4, 1),'P','1')
"""

select4 = """
select LogoSap, TypSap, Opis from W_FK7_GNIAZDA_HA ha(NOLOCK)
where rodzaj='HA' and rok=YEAR(getdate()) and LogoSap is not NULL and IsNull(TypSap, '') <> ''
order by TypSap, LogoSap
"""

select5 = """
select rp.gniazdo,  Data, 
PD_RP_H5, PD_RP_M5, PD_RP_KomentarzDoPrzestoju
from dbo.w_mg_vv_Nag_RP_E rp(NOLOCK)
join dbo.mg_vv_Lin_RejAll_E lin WITH (NOLOCK) on lin.NagId = rp.NagId
where Data >='%s' and Data<= '%s' and rd = 'PD' and pd_rp_h5 + pd_rp_M5 <> 0
group by Data, rp.gniazdo, PD_RP_H5, PD_RP_M5, PD_RP_KomentarzDoPrzestoju
order by gniazdo, Data
"""

select6 = """
select gniazdo, sum(MinProd) as MinProd, sum(MinRemont) as MinRemont, sum(MinProdNetto) as MinProdNetto
from
(
    select nag.gniazdo,
    -1*sum((nag.PD_RP_O1+nag.PD_RP_O2+nag.PD_RP_O3+nag.PD_RP_O4+nag.PD_RP_O6+nag.PD_RP_O7)*(nag.PD_RP_H1*60+nag.PD_RP_M1)/nag.pd_IloscZmian) as MinProd,
    sum((nag.PD_RP_O1+nag.PD_RP_O2+nag.PD_RP_O3+nag.PD_RP_O4+nag.PD_RP_O6+nag.PD_RP_O7)*(nag.PD_RP_H1*60+nag.PD_RP_M1)/nag.pd_IloscZmian) as MinRemont,
    -1*sum(nag.PD_RP_H2*60+nag.PD_RP_M2) as MinProdNetto
    from dbo.w_mg_vv_Nag_RP_E nag(NOLOCK)
    left join dbo.mg_vv_Lin_RejAll_E lin WITH (NOLOCK) on lin.NagId = nag.NagId
    where nag.Data >='%s' and nag.Data< '%s' and nag.rd = 'PD' and lin.NagId is NULL
    group by nag.gniazdo
    union all     
    select nag.gniazdo,
    sum((nag.PD_RP_O1+nag.PD_RP_O2+nag.PD_RP_O3+nag.PD_RP_O4+nag.PD_RP_O6+nag.PD_RP_O7)*(nag.PD_RP_H1*60+nag.PD_RP_M1)/nag.pd_IloscZmian) as MinProd,
    0 as MinRemont,
    sum(nag.PD_RP_H2*60+nag.PD_RP_M2) as MinProdNetto
    from dbo.w_mg_vv_Nag_RP_E nag(NOLOCK)
    where nag.Data >='%s' and nag.Data< '%s' and nag.rd = 'PD' 
    group by nag.gniazdo
) tab    
group by gniazdo
"""

select7 = """
select lokalizacja, Osoba1, Osoba2, Osoba3 from dbo.W_VV_MG_KIEROWNICYWYMIAROWBRO v(NOLOCK)
where Dzial = 'PR'
"""

MAIL_CONTENT = """
Plik z wyliczoną premią generowany automatycznie. Proszę o jego dokładne sprawdzenie.
"""

#MAIL_CONTENT = """
#Witam,
#w załączeniu testowy plik do wyliczania premii generowany automatycznie za miesiąc marzec 2017 r. Proszę o jego powtórne sprawdzenie. 
#Dołączam też poprawiony plik z danymi wejściowymi do wyliczania premii. W następnych miesiącach proszę o wypełnianie załączonego pliku i odsyłanie go bez zmiany nazwy pliku (ważne jest aby nazwa zaczynała się od 2 cyfrowego nr oddziału). 
#
#Pozdrawiam
#Sławomir Chołaj
#"""


def read_int(value_str):    
    if value_str:
        if type(value_str)==str:
            if value_str[0]=='=':
                try:
                    ret = int(eval(value_str[1:]))
                except:
                    ret = 0
            else:
                try:
                    ret = int(value_str)
                except:
                    ret = 0
            return ret            
        else:
            return value_str
    else:
        return 0

def read_float(value_str):
    if value_str:
        if type(value_str)==str:
            if value_str[0]=='=':
                try:
                    ret = float(eval(value_str[1:]))
                except:
                    ret = 0
            else:
                try:
                    ret = float(value_str)
                except:
                    ret = 0
            return ret            
        else:
            return value_str
    else:
        return 0

def read_str(value_str):
    if value_str:
        if type(value_str) == str:
            if value_str[0]=='=':
                try:
                    ret = eval(value_str[1:])
                except:
                    ret = 0
            else:
                try:
                    ret = value_str
                except:
                    ret = 0
            return ret            
        else:
            return value_str
    else:
        return 0


class Prac():
    def __init__(self, row=None):
        if row:
            self.nr_sap = int(row[0].value)
            self.nazwisko = row[1].value
            self.stawka = row[3].value
            self.czas = row[2].value
            self.czas_noc = row[4].value
            self.czas_nadl50 = row[5].value
            self.czas_nadl100 = row[6].value
        else:
            self.nr_sap = 0
            self.nazwisko = ""
            self.stawka = 0
            self.czas = 0
            self.czas_nadl50 = 0
            self.czas_nadl100 = 0
            self.czas_noc = 0
            

class DataOddz():
    def __init__(self):        
        self.godz_obce = 0
        self.naj_wyn = 0
        self.il_rbg = 0
        
        self.prac = {}

        
class Prod():
    def __init__(self, nag, lines, cuw, pw, awarie, godziny, data_oddz, mag):
        self.nag = nag
        self.lines = lines
        self.cuw = cuw[mag]
        self._pw = pw
        self.mag = mag
        self.nr_osob = None        
        self.data_oddz = data_oddz
        
        self.settings = models.Config.objects.filter(nr_oddz=mag)[0]
        self.gniazda = self.settings.gniazda.split(';')
        self._ilosc_gniazd =  len(self.gniazda)
        if len(self.gniazda)==1:
            self.gniazda = [self.gniazda[0], '0']
            
        self.sap_ids = sorted(self.cuw.keys())
        
        self._awarie = []
        for pos in self.gniazda:
            if pos in awarie:
                self._awarie.append(awarie[pos])
            else:
                self._awarie.append([])

        self._godziny = []
        for pos in self.gniazda:
            if pos in godziny:
                self._godziny.append(godziny[pos])
            else:
                self._godziny.append([pos, 0, 0, 0])

    def minuty_prod(self, lp_gniazda):
        if lp_gniazda < self._ilosc_gniazd:
            return self._godziny[lp_gniazda][1]
        else:
            return 0
        
    def minuty_remont(self, lp_gniazda):
        if lp_gniazda < self._ilosc_gniazd:
            return self._godziny[lp_gniazda][2]
        else:
            return 0
        
    def minuty_prod_netto(self, lp_gniazda):
        if lp_gniazda < self._ilosc_gniazd:
            return self._godziny[lp_gniazda][3]
        else:
            return 0

    def awarie(self):
        x = []
        for pos in self._awarie:
            for row in pos:
                x.append(row)
        return  x
            
    def _parametry(self, lp_gniazda, value):
        if lp_gniazda < self._ilosc_gniazd:
            v = value.split(';')
            if lp_gniazda < len(v):
                return v[lp_gniazda]
            else:
                return v[0]
        else: 
            return "0"

        
    #def wykorzystanie_maszyny(self, lp_gniazda):
    #    return self._parametry(lp_gniazda, self.settings.wykorzystanie_maszyny)
    
    #def proc_premii100(self, lp_gniazda):
    #    return self._parametry(lp_gniazda, self.settings.proc_premii_przy100)
    
    #def poziom_niepierwszy(self, lp_gniazda):
    #    return self._parametry(lp_gniazda, self.settings.poziom_niepierwszy)
    
    #def wyrob_premia(self):
    #    return self.settings.wyrob_start_premii
    
    #def korekta(self):
     #   return self.settings.korekta
    
    #def wsp(self):
    #    return self.settings.wsp
        
    def ustaw_nr_osobowy(self, nr_osob):
        if nr_osob < len(self.sap_ids):
            self.nr_osob = self.sap_ids[nr_osob]
            return self.nr_osob
        else:
            self.nr_osob  = 0
            return 0
    
    def get_prac(self, nr_osob):
        if nr_osob:
            _nr_osob = nr_osob
        else:
            _nr_osob = self.nr_osob
        
        if _nr_osob in self.cuw:
            return self.cuw[_nr_osob]
        else:
            return Prac()

    def get_prac_data(self, nr_osob):
        if nr_osob:
            _nr_osob = nr_osob
        else:
            _nr_osob = self.nr_osob
        
        if _nr_osob in self.data_oddz.prac:
            return self.data_oddz.prac[_nr_osob]
        else:
            return ["",0,0,0,0,0,0]

    
    def from_line(self, id, gniazdo, lp_kart):
        if gniazdo in self.lines:
            lines2 = self.lines[gniazdo]
            if lp_kart >= 0:
                if len(lines2) > lp_kart:
                    return lines2[lp_kart][id]
                else:
                    if id==0:
                        return ""
                    else:
                        return 0
            else:
                if id==0:
                    return "RAZEM:"
                else:
                    sum = 0
                    for pos in lines2:
                        sum+=pos[id]
                    return sum
        else:
            if id==0:
                return ""
            else:
                return 0
    
    def pw(self, gniazdo_lp, gatunek):
        gniazdo = self.gniazdo(gniazdo_lp)
        key = str(gniazdo)+"_" + str(gatunek)
        if key in self._pw:
            return self._pw[key]
        else:
            return 0
            
    def from_head(self, id, gniazdo):
        if gniazdo in self.nag:
            rec = self.nag[gniazdo]
            if len(rec)>id:
                return rec[id]
        return 0
        
    def gniazdo(self,  gniazdo_lp):
        if int(gniazdo_lp)  < len(self.gniazda):
            return self.gniazda[int(gniazdo_lp)]
        else:
            return 0
        
    def symkar(self, gniazdo, lp_kart=-1):  #IloscMGPrzel        
        return self.from_line(0, self.gniazdo(gniazdo), lp_kart)

    def ilosc_mg_przel(self, gniazdo, lp_kart=-1):  #IloscMGPrzel        
        return self.from_line(1, self.gniazdo(gniazdo), lp_kart)
        
    def ilosc_cykli(self, gniazdo, lp_kart=-1): #IloscCykli                              
        return self.from_line(2, self.gniazdo(gniazdo), lp_kart)

    def czas_cyklu(self, gniazdo, lp_kart=-1): #CzasCyklu                               
        return self.from_line(3, self.gniazdo(gniazdo), lp_kart)

    def czas_teoretyczny(self, gniazdo, lp_kart=-1): #CzasTeoretyczny
        return self.from_line(4, self.gniazdo(gniazdo), lp_kart)

    def czas_kalkulacyjny(self, gniazdo, lp_kart=-1): #CzasKalkulacyjny
        return self.from_line(5, self.gniazdo(gniazdo), lp_kart)



    def czas_pracy_minuty(self, gniazdo): # MIN_godziny_pracy 
        return self.from_head(0, self.gniazdo(gniazdo))

    def praca_netto_minuty(self, gniazdo):  #MIN_praca_netto 
        return self.from_head(1, self.gniazdo(gniazdo))

    def naprawa_minuty(self, gniazdo):  #MIN_naprawa 
        return self.from_head(2, self.gniazdo(gniazdo))

    def konserwacja_minuty(self, gniazdo):  #MIN_konswerwacja 
        return self.from_head(3, self.gniazdo(gniazdo))

    def przestoj_minuty(self, gniazdo):  #MIN_przestoj 
        return self.from_head(4, self.gniazdo(gniazdo))

    def zmiana_formy_minuty(self, gniazdo):  #MIN_zmiana_formy 
        return self.from_head(5, self.gniazdo(gniazdo))

    def brak_betonu_minuty(self, gniazdo):  #MIN_brak_betonu 
        return self.from_head(6, self.gniazdo(gniazdo))

    def brak_palet_minuty(self, gniazdo):  #MIN_brak_palet 
        return self.from_head(7, self.gniazdo(gniazdo))

    def pelna_wieza_minuty(self, gniazdo):  #MIN_pelna_wieza 
        return self.from_head(8, self.gniazdo(gniazdo))

    def ilosc_wymian_formy(self, gniazdo): #IloscWymianFormy 
        return self.from_head(9, self.gniazdo(gniazdo))

    def czas_normatywny_minuty(self, gniazdo):  #CzasNormatywny_MM                       
        return self.from_head(10, self.gniazdo(gniazdo))

    def ilosc_zmian(self, gniazdo):  #IloscZmian             
        return self.from_head(11, self.gniazdo(gniazdo))

    def obsada_prod(self, gniazdo):  #RP_01       Produkcja
        return self.from_head(12, self.gniazdo(gniazdo))

    def obsada_pozost(self, gniazdo):  #RP_02       Pozostałe
        return self.from_head(13, self.gniazdo(gniazdo))

    def obsada_magazyny(self, gniazdo):  #RP_03       Magazyny
        return self.from_head(14, self.gniazdo(gniazdo))

    def obsada_remonty(self, gniazdo):  #RP_04       Remonty
        return self.from_head(15, self.gniazdo(gniazdo))

    def obsada_zewn(self, gniazdo):  #RP_05       Zewnętrzni
        return self.from_head(16, self.gniazdo(gniazdo))

    def obsada_urlopy(self, gniazdo):  #RP_06       Urlopy
        return self.from_head(17, self.gniazdo(gniazdo))

    def obsada_zwolnienie(self, gniazdo):  #RP_07       Zwolnieni
        return self.from_head(18, self.gniazdo(gniazdo))

    def czas_kalkulacyjny_minuty(self, gniazdo):  #CzasKalkulacyjny_MM 
        return self.from_head(19, self.gniazdo(gniazdo))

    def czas_normatywny_sek(self, gniazdo):  #CzasNormatywny
        return self.from_head(20, self.gniazdo(gniazdo))

    # CZASY

    def nazwisko(self,  nr_osob=None):
        return self.get_prac(nr_osob).nazwisko

    def stawka(self, nr_osob=None):
        return self.get_prac(nr_osob).stawka

    def godz_przepr(self, nr_osob=None):
        return self.get_prac(nr_osob).czas

    def nadl_50(self, nr_osob=None):
        return self.get_prac(nr_osob).czas_nadl50
    
    def nadl_100(self, nr_osob=None):
        return self.get_prac(nr_osob).czas_nadl100

    def godz_nocne(self, nr_osob=None):
        return self.get_prac(nr_osob).czas_noc

    def oddzial(self):
        return self.settings.nazwa

    ##
        
    def najnizsze(self):
        return self.data_oddz.naj_wyn
    
    def ilosc_rbg(self):
        return self.data_oddz.il_rbg
         
    def godz_obce(self):
        return self.data_oddz.godz_obce

    ##
    def punkty_za_prod(self,  nr_osob=None):
        return self.get_prac_data(nr_osob)[1]

    def potracenia(self,  nr_osob=None):
        return self.get_prac_data(nr_osob)[2]

    def nagrody(self,  nr_osob=None):
        return self.get_prac_data(nr_osob)[3]

    def dodatek_bryg(self,  nr_osob=None):
        return self.get_prac_data(nr_osob)[4]

    def dodatek_bhp(self,  nr_osob=None):
        return self.get_prac_data(nr_osob)[5]

    def pranie(self,  nr_osob=None):
        return self.get_prac_data(nr_osob)[6]

    ##
    def numer_lokalizacji(self):
        return self.mag

    def _replace_proc(self, value):
        if '%' in value:
            return float(value.replace('%', ''))/100
        else:
            return float(value)

    ##
    def p1(self):
        return self._replace_proc(self.settings.p1)

    def p2(self):
        return self._replace_proc(self.settings.p2)

    def p3(self):
        return self._replace_proc(self.settings.p3)

    def p4(self):
        return self._replace_proc(self.settings.p4)

    def p5(self):
        return self._replace_proc(self.settings.p5)

    def p6(self):
        return self._replace_proc(self.settings.p6)

    def p7(self):
        return self._replace_proc(self.settings.p7)

    def p8(self):
        return self._replace_proc(self.settings.p8)
        
        
        
def get_oddz_from_name(oddz):
    conf = models.Config.objects.all()
    x= oddz[:3].upper()
    for  pos in conf:
        if pos.nazwa[:3].upper()==x:
            return "%02d" % pos.nr_oddz
    return '50'
     

PFORM = form_with_perms('produkcja') 


class ProdukcjaForm(forms.Form):
    rok = forms.ChoiceField(label=_('Rok'), required=True, initial='2016',choices=models.Rok_CHOICES)
    mies = forms.ChoiceField(label=_('Miesiąc'), required=True, initial='01',choices=models.Miesiac_CHOICES)
    ods_wzr = forms.FileField(label=_('Plik[i] z oddziału[ów]'), required=False, )
    stawki = forms.FileField(label=_('Stawki i czasy (CUW)'), required=False, )
    test = forms.BooleanField(label=_('Test'), required=False, initial=True,)
    
    def process(self, request, queryset=None):
    
        LIN_LIMIT  = 10
        proc_id = int(request.POST.get('proc_id', '0'))
        
        tmp_dir = gettempdir()
        
        if proc_id == 0:
            #try:
                odfdata= request.FILES['ods_wzr']
                stawki = request.FILES['stawki']
        
                stream = io.BytesIO(odfdata.read())
                zfile = zipfile.ZipFile(stream)
        
                for m in zfile.infolist():
                    data = zfile.read(m) 
                    disk_file_name = os.path.join(tmp_dir, m.filename[:2]+'.xlsx')
                    with open(disk_file_name, 'wb') as fd:
                        fd.write(data)
        
                file_name = os.path.join(tmp_dir,'temp_stawki.xlsx')
                with open(file_name, 'wb') as f:
                    f.write(stawki.read())
        
                return { 'status':  'start', 'OK': True }
            #except:
            #    return { 'status':  'start', 'OK': False, 'message': sys.exc_info() }
        
        elif proc_id == 1000:    
            try:                
                file_list = os.listdir(path=tmp_dir)
                for pos in file_list:
                    if len(pos)>3 and pos[0]>='0' and pos[0]<='9' and pos[1]>='0' and pos[1]<='9' and pos.lower().endswith('.xlsx'):
                        file_name = os.path.join(tmp_dir, pos)
                        #os.unlink(file_name)
                file_name = os.path.join(tmp_dir, 'temp_stawki.xlsx')
                os.unlink(file_name)
                return { 'status':  'end', 'OK': True }
            except:
                return { 'status':  'end', 'OK': False, 'message': sys.exc_info()  }
        
        elif proc_id > 0 and proc_id < 1000:
            print(proc_id)
            
            rok = self.cleaned_data['rok']
            mies = self.cleaned_data['mies']
        
            if int(mies) == 1:
                data_start = "%04d%02d28" % ( int(rok)-1, 12)
                data_stop = "%04d%02d28" % ( int(rok), 1)
                data_start_str = "%04d-%02d-28" % ( int(rok)-1, 12)
                data_stop_str = "%04d-%02d-27" % ( int(rok), 1)
            else:
                data_start = "%04d%02d28" % ( int(rok), int(mies)-1)
                data_stop = "%04d%02d28" % ( int(rok), int(mies))
                data_start_str = "%04d-%02d-28" % ( int(rok), int(mies)-1)
                data_stop_str = "%04d-%02d-27" % ( int(rok), int(mies))
                
            test = self.cleaned_data['test']
        
            softlab = {}
        
        
            file_list = os.listdir(path=tmp_dir)
            data_path = None
            for pos in file_list:
                if len(pos)>3 and pos[0]>='0' and pos[0]<='9' and pos[1]>='0' and pos[1]<='9' and pos.lower().endswith('.xlsx'):
                    x = int(pos[:2])
                    if x == proc_id:
                        #data_path = pos
                        data_path = os.path.join(tmp_dir, pos)
        
            #data_path = os.path.join(tmp_dir, "Dane_prem_%02d.xlsx" % proc_id)
            #print(data_path)
            if not data_path or not os.path.exists(data_path):                
                return { 'status':  'etap', 'proc_id': proc_id, 'OK': False, 'message': 'Brak danych' }
            
            workbook = openpyxl.load_workbook(filename=data_path, read_only=True)
            worksheets = workbook.get_sheet_names()
            worksheet = workbook.get_sheet_by_name(worksheets[0])
            i = 0
                                
            data = DataOddz()
            
            for row in worksheet.rows:
                if i==1:
                    data.godz_obce = read_int(row[1].value)
                elif i==2:
                    data.naj_wyn = read_int(row[1].value)
                elif i==3:
                    data.il_rbg = read_int(row[1].value)
                elif i<8:
                    pass
                else:
                    try:
                        print("LP:", row[1].value)
                        no =int(row[1].value)
                    except:
                        no = -1
                    if no>1000 and no<100000:
                        try:
                            r = []
                            for i in range(2,9):
                                r.append(read_str(row[i].value))
                                #if row[i].value:
                                #    r.append(row[i].value)
                                #else:
                                #   r.append(0)                                
                            data.prac[int(row[1].value)] = r
                        except:
                            pass
                i+=1
                
            workbook._archive.close()
            #print(data.prac)
            maile = {}
            with settings.DB as db:
                db.execute(select7)
                for pos in db.fetchall():
                    maile[int(pos[0])] = (pos[1], pos[2], pos[3],)
                
                #db.execute(select4)
                #for pos in db.fetchall():
                #    try:
                #        nr = int(pos[0])
                #        oddz = pos[1][-2:]
                #        opis = pos[2]
                #        softlab[nr] = (oddz, opis)
                #    except:
                #        continue
        
                cuw = {}
                file_name = os.path.join(tmp_dir, "temp_stawki.xlsx")
                print(file_name)
                workbook = openpyxl.load_workbook(filename=file_name, read_only=True)
                worksheets = workbook.get_sheet_names()
                print(worksheets)
                worksheet = workbook.get_sheet_by_name(worksheets[0])
                print(worksheet)
                for row in worksheet.rows:
                    try:
                        sap_id = int(row[0].value)
                    except:
                        continue
                    if sap_id <= 0:
                        continue
                    prac = (get_oddz_from_name(row[7].value), row[1].value)
                    softlab[sap_id] = prac
                    if not int(prac[0]) in cuw:
                        cuw[int(prac[0])] = {}
                    cuw[int(prac[0])][sap_id] = Prac(row)
                
                workbook._archive.close()
                
                nag = {}
                db.execute(select1 % (data_start, data_stop))        
                for pos in db.fetchall():
                    nag[pos[0]]=pos[1:]
        
                lin = {}
                db.execute(select2 % (data_start, data_stop))        
                for pos in db.fetchall():
                    if pos[0] in lin:
                        lin[pos[0]].append(pos[1:])
                    else:
                        lin[pos[0]] = [pos[1:],]
                    
                for key, tab in lin.items():  
                    if len(tab)>LIN_LIMIT:
                        i = len(tab)
                        row = None
                        while i>LIN_LIMIT:
                            if row:
                                for j in range(len(tab[i-1])):
                                    if j==0 or j==3:
                                        pass
                                    else:
                                        row[j] += tab[i-1][j]
                            else:
                                row = list(tab[i-1])
                                row[0]='POZOSTALE:'
                                row[3] = row[4]/row[2]
                            i-=1
                        del tab[LIN_LIMIT:]
                        tab.append(row)
        
                pw = {}
                db.execute(select3 % (data_start, data_stop))        
                for pos in db.fetchall():
                    pw[str(pos[0])+"_"+str(pos[1])] = pos[2]
                
                awarie = {}
                db.execute(select5 % (data_start, data_stop)) 
                for pos in db.fetchall():
                    gniazdo = str(pos[0])
                    if not gniazdo in awarie:
                        awarie[gniazdo] = []
                    awarie[gniazdo].append(pos)
        
        
                godziny = {}
                db.execute(select6 % (data_start, data_stop, data_start, data_stop)) 
                for pos in db.fetchall():
                    gniazdo = str(pos[0])
                    godziny[gniazdo]=pos
        
            doc_type = 'odf'
            prod = Prod(nag, lin, cuw, pw, awarie, godziny, data,  proc_id)
        
            dfile = os.path.join(settings.DATA_PATH, settings.APPSET_NAME)
            kalkulacja = os.path.join(dfile, "kalkulacja premii.ods")
            file_out, file_in = render_odf(kalkulacja, Context({ 'prod': prod, 'data_start': data_start_str, 'data_stop': data_stop_str,  }))
        
            new_name = os.path.join(tmp_dir, "%02d - premia.ods" % proc_id)
            os.rename(file_out, new_name)
        
            if True:
                with open(new_name, "rb") as f:                    
                    with open(data_path, "rb") as f2:
                        if test:
                            mail_to = [ 'slawomir.cholaj@polbruk.pl', 'tomasz.kulis@polbruk.pl', ]
                            #mail_to = [ 'slawomir.cholaj@polbruk.pl', ]
                        else:
                            mail_to = ["artur.sackiewicz@polbruk.pl", maile[proc_id][1]+"@polbruk.pl", maile[proc_id][0]+"@polbruk.pl",  'premia@polbruk.pl', 'slawomir.cholaj@polbruk.pl' ]
                                                
                        mail = EmailMessage("Plik premiowy dla oddziału: "+prod.settings.nazwa , MAIL_CONTENT, to=mail_to)
                        mail.attach(prod.settings.nazwa + "_"+data_stop.replace('-','').replace(' ','')+".ods", f.read(), "application/vnd.oasis.opendocument.spreadsheet")
                        mail.attach(("%02d_" % proc_id) + prod.settings.nazwa + ".xlsx", f2.read(), "application/vnd.ms-excel")                
                        mail.send()                
                os.unlink(new_name)
                os.unlink(data_path)
            return { 'status':  'etap', 'proc_id': proc_id, 'OK': True, 'file_name': file_out }
        else:    
            return {}
    
    def process_empty(self, request, param=None):
        return { 'config':  models.Config.objects.all() }
        

def view_produkcjaform(request, *argi, **argv):
    return PFORM(request, ProdukcjaForm, 'produkcja/formprodukcjaform.html', {})


class ConfigUploadForm(forms.Form):
    config_file = forms.FileField(label=_('Config file'), required=True, )
    
    def process(self, request, queryset=None):
    
        def format_param_proc(param):
            ret = "%.8f" % float(param*100)
            return ret + "%"
            
        def format_param(param):
            ret = "%.8f" % float(param)
            return ret
            
        config = request.FILES['config_file']
        tmp_dir = gettempdir()
        
        file_name = os.path.join(tmp_dir,'temp_config.xlsx')
        with open(file_name, 'wb') as f:
            f.write(config.read())
        
        workbook = openpyxl.load_workbook(filename=file_name, read_only=True)
        worksheets = workbook.get_sheet_names()
        worksheet = workbook.get_sheet_by_name(worksheets[0])
        for row in worksheet.rows:
            try:
                mag = int(row[0].value)
            except:
                continue
            if not (mag >= 50 and mag<=99):
                continue
            
            objects = models.Config.objects.filter(nr_oddz=mag)
            if len(objects) == 1:
                obj = objects[0]
                obj.p1 = format_param_proc(row[1].value)
                obj.p2 = format_param_proc(row[2].value)
                obj.p3 = format_param_proc(row[3].value)
                obj.p4 = format_param_proc(row[4].value)
                obj.p5 = format_param_proc(row[5].value)
                obj.p6 = format_param_proc(row[6].value)
                obj.p7 = format_param_proc(row[7].value)
                obj.p8 = format_param(row[8].value)
                obj.save()
        workbook._archive.close()
        
        return { 'OK': True }
    

def view_configuploadform(request, *argi, **argv):
    return PFORM(request, ConfigUploadForm, 'produkcja/formconfiguploadform.html', {})




 
